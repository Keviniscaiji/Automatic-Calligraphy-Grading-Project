# 生成切割后的labels
import os.path

import openpyxl
import pandas
import xlrd as xlrd
from openpyxl import load_workbook

from load_data import *

base_path = os.path.join("dataset")


def remove_zeros(label_path):
    label_path = os.path.join(base_path, label_path)
    error_list = []
    for file_path in os.listdir(label_path):
        data = xlrd.open_workbook(os.path.join(label_path, file_path))
        xl = pandas.ExcelFile(os.path.join(label_path, file_path))
        sheet_names = xl.sheet_names
        for sheet_name in sheet_names:
            table = data.sheet_by_name(sheet_name)
            nrows = table.nrows
            ncols = table.ncols
            for i in range(0, nrows):
                row = table.row_values(i)
                for j in range(0, ncols):
                    item = row[j]
                    if is_number(item):
                        if item < 0:
                            print(file_path, sheet_name, str(i), str(j))
                            error_list.append([file_path, sheet_name, str(i), str(j)])
    print(error_list)
    # for item in error_list:
    #     workbook = openpyxl.load_workbook(os.path.join(label_path, item[0]))
    #     worksheet = workbook.worksheets[0]
    #     worksheet.cell(int(item[2]), int(item[3]), 0)
    #     workbook.save(filename=os.path.join(label_path, item[0]))
    # for item in error_list:


def cut_excel(pic_line_num, restored_parameter_num, label_path, cut_labels_path):
    label_path = os.path.join(base_path, label_path)
    cut_labels_path = os.path.join(base_path, cut_labels_path)
    if not os.path.exists(label_path):
        os.mkdir(label_path)
    if not os.path.exists(cut_labels_path):
        os.mkdir(cut_labels_path)

    rownumAndTotalscore = {}

    for path in os.listdir(label_path):
        wb = xlrd.open_workbook(os.path.join(label_path, path))
        # score_list = []
        row_num_size = 0
        for sh in wb.sheets():
            score_all_col = [[] for i in range(restored_parameter_num)]  # 5 种分数的列表，每个列表12/14行
            for i in range(1, sh.nrows):
                row = sh.row_values(i)
                if is_number(row[0]):
                    index = int(row[0])
                if is_number(row[2]):
                    row_num_size += 1

                    print(row_num_size)

                    # 统计每行的得分
                    score_all_col[0].append(max(float(row[11]), 0)) if is_number(row[11]) else score_all_col[0].append(
                        0)
                    score_all_col[1].append(max(float(row[13]), 0)) if is_number(row[13]) else score_all_col[1].append(
                        0)
                    for j in range(2, restored_parameter_num):
                        score_all_col[j].append(max(float(row[13 + j]), 0)) if is_number(row[13 + j]) else \
                            score_all_col[j].append(0)

                if row[2] == '错误总数：':
                    col_num = len(score_all_col[0])
                    row_num = 1
                    for j in range(col_num - row_num_size, col_num - pic_line_num + 1):
                        # print(col_num - row_num_size)
                        # col_num：所有数据的行数
                        # row_num_size：当前数据所包含的行数
                        # pic_line_num：切割后的图片每张图片所包含的行数
                        # col_num-row_num_size+(row_num_size-pic_line_num)+1 化简后得到上边界 col_num-pic_line_num+1
                        score_list = []
                        for score_index in range(restored_parameter_num):
                            score_list.append(np.mean(score_all_col[score_index][j:j + pic_line_num]))

                        #     计算line_num行的平均值
                        f = open(os.path.join(cut_labels_path, '{}-{}-{}.txt'.format(index, pic_line_num, row_num)),
                                 'w')
                        f.write('\n'.join(list(map(str, score_list))))
                        f.close()
                        row_num += 1
                        row_num_size = 0

def cut_excel_finalscore(label_path, cut_labels_path):
    label_path = os.path.join(base_path, label_path)
    cut_labels_path = os.path.join(base_path, cut_labels_path)
    if not os.path.exists(label_path):
        os.mkdir(label_path)
    if not os.path.exists(cut_labels_path):
        os.mkdir(cut_labels_path)

    for path in os.listdir(label_path):
        wb = xlrd.open_workbook(os.path.join(label_path, path))
        # score_list = []
        row_num_size = 0
        for sh in wb.sheets():
            # score_all_col = [[] for i in range(restored_parameter_num)]  # 5 种分数的列表，每个列表12/14行
            for i in range(1, sh.nrows):
                row = sh.row_values(i)
                # 获得当前读取的文档的编号
                if is_number(row[0]):
                    index = int(row[0])
                #  获得当前读取文档的
                # if is_number(row[2]):
                #     row_num_size += 1
                #     score_all_col[0].append(max(float(row[11]), 0)) if is_number(row[11]) else score_all_col[0].append(
                #         0)
                #     score_all_col[1].append(max(float(row[13]), 0)) if is_number(row[13]) else score_all_col[1].append(
                #         0)
                #     for j in range(2, restored_parameter_num):
                #         score_all_col[j].append(max(float(row[13 + j]), 0)) if is_number(row[13 + j]) else \
                #             score_all_col[j].append(0)
                if row[2] == '错误总数：':
                    # col_num = len(score_all_col[0])
                    # row_num = 1
                    finalScoreRowNum = i + 1
                    newRow = sh.row_values(finalScoreRowNum)
                    # for j in range(col_num - row_num_size, col_num - pic_line_num + 1):
                    #     # print(col_num - row_num_size)
                    #     # col_num：所有数据的行数
                    #     # row_num_size：当前数据所包含的行数
                    #     # pic_line_num：切割后的图片每张图片所包含的行数
                    #     # col_num-row_num_size+(row_num_size-pic_line_num)+1 化简后得到上边界 col_num-pic_line_num+1
                    #     score_list = []
                    #     for score_index in range(restored_parameter_num):
                    #         score_list.append(np.mean(score_all_col[score_index][j:j + pic_line_num]))
                    final_score = newRow[20]
                        #     计算line_num行的平均值
                    f = open(os.path.join(cut_labels_path, '{}.txt'.format(index)), 'w')
                    f.write(str(final_score))
                    f.close()
                        # row_num += 1
                        # row_num_size = 0

def cut_excel_overlap_patch(pic_line_num, restored_parameter_num, label_path, cut_labels_path):
    label_path = os.path.join(base_path, label_path)
    cut_labels_path = os.path.join(base_path, cut_labels_path)
    if not os.path.exists(label_path):
        os.mkdir(label_path)
    if not os.path.exists(cut_labels_path):
        os.mkdir(cut_labels_path)

    for path in os.listdir(label_path):
        print(path)
        wb = xlrd.open_workbook(os.path.join(label_path, path))
        # score_list = []
        row_num_size = 0
        for sh in wb.sheets():
            score_all_col = [[] for i in range(restored_parameter_num)]  # 5 种分数的列表，每个列表12/14行
            for i in range(1, sh.nrows):
                row = sh.row_values(i)
                if is_number(row[0]):
                    index = int(row[0])
                if is_number(row[2]):
                    row_num_size += 1
                    score_all_col[0].append(max(float(row[11]), 0)) if is_number(row[11]) else score_all_col[0].append(
                        0)
                    score_all_col[1].append(max(float(row[13]), 0)) if is_number(row[13]) else score_all_col[1].append(
                        0)
                    for j in range(2, restored_parameter_num):
                        score_all_col[j].append(max(float(row[13 + j]), 0)) if is_number(row[13 + j]) else \
                            score_all_col[j].append(0)
                if row[2] == '错误总数：':
                    col_num = len(score_all_col[0])
                    row_num = 1

                    for j in range(col_num - row_num_size, col_num):
                        # print(col_num - row_num_size)
                        # col_num：所有数据的行数
                        # row_num_size：当前数据所包含的行数
                        # pic_line_num：切割后的图片每张图片所包含的行数
                        # col_num-row_num_size+(row_num_size-pic_line_num)+1 化简后得到上边界 col_num-pic_line_num+1
                        if j <= col_num - pic_line_num:
                            score_list = []
                            for score_index in range(restored_parameter_num):
                                score_list.append(np.mean(score_all_col[score_index][j:j + pic_line_num]))
                            f = open(os.path.join(cut_labels_path, '{}-{}-{}.txt'.format(index, pic_line_num, row_num)),
                                     'w')
                            f.write('\n'.join(list(map(str, score_list))))
                            f.close()
                            row_num += 1

                            #     计算line_num行的平均值
                        elif j <= col_num:
                            cut_list_indexs = []
                            score_list = []
                            cut_list = []
                            difference = pic_line_num - (col_num - j)

                            for i in range(j, col_num):
                                cut_list_indexs.append(i)
                            for i in range(col_num - row_num_size, col_num - row_num_size + difference):
                                cut_list_indexs.append(i)
                            for score_index in range(restored_parameter_num):
                                for cut_list_index in cut_list_indexs:
                                    cut_list.append(score_all_col[score_index][cut_list_index])
                                score_list.append(np.mean(cut_list))
                            print(cut_list_indexs)
                            f = open(os.path.join(cut_labels_path, '{}-{}-{}.txt'.format(index, pic_line_num, row_num)),
                                     'w')
                            f.write('\n'.join(list(map(str, score_list))))
                            f.close()
                            row_num += 1

                            #     计算line_num行的平均值



                    row_num_size = 0


def cut_excel_without_overlap(pic_num, restored_parameter_num, label_path, cut_labels_path):
    label_path = os.path.join(base_path, label_path)
    cut_labels_path = os.path.join(base_path, cut_labels_path)
    if not os.path.exists(label_path):
        os.mkdir(label_path)
    if not os.path.exists(cut_labels_path):
        os.mkdir(cut_labels_path)

    for path in os.listdir(label_path):
        wb = xlrd.open_workbook(os.path.join(label_path, path))
        # score_list = []
        row_num_size = 0
        for sh in wb.sheets():
            score_all_col = [[] for i in range(restored_parameter_num)]  # 7 种分数的列表，每个列表12/14行
            for i in range(1, sh.nrows):
                row = sh.row_values(i)
                if is_number(row[0]):
                    index = int(row[0])
                #     保存文件的标签
                if is_number(row[2]) and is_number(row[3]) and not is_number(row[20]):
                    row_num_size += 1
                    score_all_col[0].append(max(float(row[11]), 0)) if is_number(row[11]) else score_all_col[0].append(
                        0)
                    score_all_col[1].append(max(float(row[13]), 0)) if is_number(row[13]) else score_all_col[1].append(
                        0)
                    for j in range(2, restored_parameter_num):
                        score_all_col[j].append(max(float(row[13 + j]), 0)) if is_number(row[13 + j]) else \
                            score_all_col[j].append(0)
                if row[2] == '错误总数：':
                    total_row_num = len(score_all_col[0])
                    row_num = 1
                    # print(index)
                    k = 0
                    while k < row_num_size - pic_num + 1:
                        # while k <1:
                        # col_num：所有数据的行数
                        # row_num_size：当前数据所包含的行数
                        # pic_line_num：切割后的图片每张图片所包含的行数
                        score_list = []
                        for score_index in range(restored_parameter_num):
                            score_list.append((np.mean(score_all_col[score_index][
                                                       total_row_num + k - row_num_size:max(
                                                           total_row_num + k - row_num_size + pic_num,
                                                           total_row_num)])))
                            # score_list.append((np.mean(score_all_col[score_index][total_row_num + k - row_num_size:total_row_num])))

                            print(str(row_num_size), str(total_row_num), str(k))
                        #     计算line_num行的平均值
                        f = open(os.path.join(cut_labels_path, '{}-{}-{}.txt'.format(index, pic_num, row_num)),
                                 'w')
                        # f = open(os.path.join(img_new_path, '{}.txt'.format(index)),
                        #          'w')
                        f.write('\n'.join(list(map(str, score_list))))
                        f.close()
                        row_num += 1
                        k += 1
                        # score_all_col = [[] for i in range(restored_parameter_num)]
                    row_num_size = 0


if __name__ == '__main__':
    # remove_zeros("marked_forms")
    # cut_excel_finalscore("marked_forms", "cut_labels")
    # cut_excel_overlap_patch(3, 5, "marked_forms", "cut_labels")
    cut_excel(1,7, "marked_forms","labelsLine")
    # cut_excel(1,5, "marked_forms", "labels_5parameters_1_lines_overlap")
    # cut_excel_without_overlap(1,7, "marked_forms","labels_7parameters_1_lines_no_overlap")
    # cut_excel_without_overlap(1,5, "marked_forms", "labels_5parameters_1_lines_no_overlap")
    # cut_excel(2,7, "marked_forms","labels_7parameters_2_lines_overlap")
    # cut_excel(2,5, "marked_forms", "labels_5parameters_2_lines_overlap")
    # cut_excel_without_overlap(2,7, "marked_forms","labels_7parameters_2_lines_no_overlap")
    # cut_excel_without_overlap(2,5, "marked_forms", "labels_5parameters_2_lines_no_overlap")
    # cut_excel(3,7, "marked_forms","labels_7parameters_3_lines_overlap")
    # cut_excel(3,5, "marked_forms", "labels_5parameters_3_lines_overlap")
    # cut_excel(4,7, "marked_forms","labels_7parameters_4_lines_overlap")
    # cut_excel(4,5, "marked_forms", "labels_5parameters_4_lines_overlap")
    # cut_excel(5,7, "marked_forms","labels_7parameters_5_lines_overlap")
    # cut_excel(5,5, "marked_forms", "labels_5parameters_5_lines_overlap")
    # cut_excel(1,5,"marked_forms",os.path.join("2022_10_19dataset","labels_1line_patch"))
    # cut_excel_overlap_patch(3, 5, "marked_forms","dataset/2022_11_2dataset/labels_7parameters_3_line_overlap_patch")
    # cut_excel_overlap_patch(3, 5, "labels_7parameters_3_line_overlap_patch")
    # cut_excel_overlap_patch(3, 7, "labels_7parameters_3_line_overlap_patch")
    # cut_excel(12,7, "marked_forms","labels_7parameters_12_line_no_overlap")
    # cut_excel(12,5, "marked_forms", "labels_5parameters_12_lines_no_overlap")
    # cut_excel_whole(1,7, "marked_forms","labels_7parameters_whole_lines_no_overlap")
    # cut_excel_whole(1,5, "marked_forms", "labels_5parameters_whole_lines_no_overlap")
    # cut_excel_without_overlap(12, 5, "marked_forms", "labels_5parameters_whole_pic")
    # cut_excel_without_overlap(12, 7, "marked_forms", "labels_7parameters_whole_pic")
